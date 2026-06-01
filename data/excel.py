import re
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from time import sleep
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from public import shared_data
from public.jiaoyisuo import select_future_dict
from public.print_context import print_context


def init_trade_excel():
    base_dir = os.getcwd()
    save_dir = os.path.join(base_dir, r"trade_data")
    os.makedirs(save_dir, exist_ok=True)
    filename = os.path.join(save_dir, "trade.xlsx")

    if not os.path.exists(filename):
        wb = Workbook()
        ws_trade = wb.active
        ws_trade.title = "交易记录"
        wb.create_sheet("统计表")

        headers = [
            "建仓交易时间", "平仓交易时间", "用户", "期货品种", "交易方向", "交易价",
            "止损价", "止盈价", "交易前总额", "交易后总额", "利润"
        ]

        # 标题（合并行）
        end_col = get_column_letter(len(headers))
        ws_trade.merge_cells(f"A1:{end_col}1")
        title_cell = ws_trade["A1"]
        title_cell.value = "期货交易明细表"
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=16)

        # 写入表头
        ws_trade.append(headers)

        # 表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        for col_num in range(1, len(headers) + 1):
            cell = ws_trade.cell(row=2, column=col_num)
            cell.font = header_font
            cell.alignment = header_alignment

        wb.save(filename)
        wb.close()

    return filename


def auto_adjust_column_width(ws):
    """
    修复版：自动适应列宽，跳过合并单元格，绝不报错
    """
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_num)

        # 遍历每一行，计算最长内容
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # 设置列宽
        ws.column_dimensions[col_letter].width = max_length + 3


def save_trade_data(data_list):  # 追加交易数据到excel工作薄工作表[交易记录]
    filename = init_trade_excel()
    wb = load_workbook(filename)
    ws = wb["交易记录"]

    # 插入到第3行，最新数据在最上面
    ws.insert_rows(3)

    # 写入数据 + 居中
    for col, value in enumerate(data_list, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自动列宽（已修复报错）
    auto_adjust_column_width(ws)
    try:
        wb.save(filename)
        wb.close()
        print(f"✅ 交易记录写入成功！")
    except:
        print_context("保存数据出错：请检查保存文件路径或文件正在使用拒绝写入！")
        return
    # print(f"📁 文件路径：{filename}")
    # print(f"📄 记录内容：{data_list}\n"..., ,,mmmm)


def monitor_trade():  # 设置成线程监控是否完成交易（平仓）
    try:
        future_code = shared_data.ts_code
        margin_ctrl = shared_data.ths_common_control['盈亏'][0]
        char = re.findall(r'[A-Za-z]+', future_code)[0]
        suffix = re.findall(r'\d+', future_code)[0]
        future_name = select_future_dict[char][0] + suffix

        while True:
            if margin_ctrl.window_text() == future_name and shared_data.is_trading_flag:
                print_context(f"✅ [{future_code}] 已平仓,正在保存数据············")
                # 将交易数据写入到指定文件：G:\auto_future\data\trade_data\trade.xlsx
                data = [shared_data.start_trade_date, datetime.now(), shared_data.user_name, shared_data.ts_code,
                        "buy", 1010, 980, 940, 1020, 142600, 160000, 17400]
                save_trade_data(data)
                shared_data.is_trading_flag = False
            else:
                print(f"🔔 [{future_code}] 持仓交易中...")
            sleep(10)
    except Exception as e:
        print(f"⚠️ 监控异常：{e}")


if __name__ == '__main__':
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    test_data = [
        now,
        now,
        "1811889418",
        "au2606",
        "Rise",
        1010,
        980,
        1024,
        1426000,
        1446238,
        20238
    ]
    save_trade_data(test_data)
